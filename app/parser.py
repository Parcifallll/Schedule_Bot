from openpyxl import load_workbook
from openpyxl.cell.rich_text import CellRichText, TextBlock
from urllib.parse import urlencode
from requests import get
from datetime import date
from re import search, findall
from json import load


def yearfunc(month):
    if int(month) < 9:
        return 2025
    else:
        return 2024


def strikethrough_clear(cell):
    strikethrough_lst = []
    if isinstance(cell.value, CellRichText):
        for text_block in cell.value:
            if isinstance(text_block, TextBlock) and text_block.font.strikethrough:
                strikethrough_lst.append(text_block.text)
    string = str(cell.value)
    for el in strikethrough_lst:
        string = string.replace(el, '')
    return string


def pars_func(subjectstring, locationstring, provided_date):
    match1 = search(r'с \d\d.\d\d.', subjectstring)
    match2 = search(r'\d\d.\d\d.-\d\d.\d\d.', subjectstring)
    match3 = findall(r'\d\d.\d\d.', subjectstring)
    match4 = search(r'ауд. \d\d\d', subjectstring)

    if match1:
        if date(yearfunc(match3[0].split('.')[1]), int(match3[0].split('.')[1]),
                int(match3[0].split('.')[0])) > provided_date:
            subjectstring = None
            locationstring = None

    elif match2:
        from_date = date(yearfunc(match3[0].split('.')[1]), int(match2[0].split('-')[0].split('.')[1]),
                         int(match2[0].split('-')[0].split('.')[0]))
        to_date = date(yearfunc(match3[0].split('.')[1]), int(match2[0].split('-')[1].split('.')[1]),
                       int(match2[0].split('-')[1].split('.')[0]))
        if not (from_date <= provided_date <= to_date):
            subjectstring = None
            locationstring = None

    elif match3 and match4:
        if 'БП' in locationstring:
            building = 'БП'
        elif 'Л' in locationstring:
            building = 'Л'
        elif 'Р' in locationstring:
            building = 'Р'
        if date(yearfunc(match3[0].split('.')[1]), int(match3[0].split('.')[1]),
                int(match3[0].split('.')[0])) == provided_date:
            locationstring = f'{match4[0][5:]} - {building}'
        match4 = match4[0]

    elif match3:
        date_lst = []
        for i in range(len(match3)):
            date_lst.append(
                date(yearfunc(match3[0].split('.')[1]), int(match3[i].split('.')[1]), int(match3[i].split('.')[0])))
        if not (provided_date in date_lst):
            subjectstring = None
            locationstring = None

    return subjectstring, locationstring


def lesson_num_to_time(lesson_number):
    match lesson_number:
        case 0:
            return '08:00 - 09:20'
        case 1:
            return '09:30 - 10:50'
        case 2:
            return '11:10 - 12:30'
        case 3:
            return '13:00 - 14:20'
        case 4:
            return '14:40 - 16:00'
        case 5:
            return '16:20 - 17:40'
        case 6:
            return '18:10 - 19:30'
        case 7:
            return '19:40 - 21:00'


def parser_main(username, datee):
    schedule_list = []

    subjects = ['Программирование на C++', 'Технология программирования', 'Дискретная математика',
                'Линейная алгебра и геометрия', 'Математический анализ', 'Безопасность жизнедеятельности',
                'История России', 'Основы российской государственности']

    with open("users.json", "r") as f:
        parsed_dict = load(f)
        group = parsed_dict[username]['group']

    provided_date = datee
    day_of_the_week = provided_date.isoweekday()

    base_url = 'https://cloud-api.yandex.net/v1/disk/public/resources/download?'
    public_key = 'https://disk.yandex.ru/i/C3VDm7Vg6LIOXA'

    final_url = base_url + urlencode(dict(public_key=public_key))
    response = get(final_url)
    download_url = response.json()['href']

    download_response = get(download_url)
    with open('excelfile.xlsx', 'wb') as ef:
        ef.write(download_response.content)

    excel_file = 'excelfile.xlsx'

    workbook = load_workbook(filename=excel_file, rich_text=True)
    worksheet = workbook.active

    min_col = 5 + 3 * (group - 1 + (group - 1) // 3)
    max_col = min_col + 1

    match day_of_the_week:
        case 1:
            min_row = 12
            max_row = 17
        case 2:
            min_row = 23
            max_row = 30
        case 3:
            min_row = 34
            max_row = 41
        case 4:
            min_row = 51
            max_row = 52
        case 5:
            min_row = 56
            max_row = 62
        case 6:
            min_row = 67
            max_row = 73

    if day_of_the_week == 4:
        l_num = 6
    else:
        l_num = 0

    for row in worksheet.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col):

        lessontime = lesson_num_to_time(l_num)
        subject = ''
        stype = ''
        location = ''

        subjectstring = strikethrough_clear(row[0])
        locationstring = strikethrough_clear(row[1])

        if ('\n---------------------\n' in subjectstring) and ('\n------\n' in locationstring):

            subjectstrings = [subjectstring.split('\n\n---------------------\n')[0],
                              subjectstring.split('\n---------------------\n')[1]]
            locationstrings = [locationstring.split('\n------\n')[0], locationstring.split('\n------\n')[1]]

            for i in range(2):
                ss, ls = pars_func(subjectstrings[i], locationstrings[i], provided_date)
                for s in subjects:
                    if s in str(ss):
                        subject = s
                        location = str(ls).replace('\n', '')
                        if 'лекция' in ss:
                            stype = 'лекция'
                        elif ('семинар' in ss) or ('практическое занятие' in ss):
                            stype = 'семинар'
                        else:
                            stype = ''

        elif ('\n\n' in subjectstring) and ('\n\n' in locationstring):

            subjectstrings = [subjectstring.split('\n\n')[0], subjectstring.split('\n\n')[1]]
            locationstrings = [locationstring.split('\n\n')[0], locationstring.split('\n\n')[1]]

            for i in range(2):
                ss, ls = pars_func(subjectstrings[i], locationstrings[i], provided_date)
                for s in subjects:
                    if s in str(ss):
                        subject = s
                        location = str(ls).replace('\n', '')
                        if 'лекция' in ss:
                            stype = 'лекция'
                        elif ('семинар' in ss) or ('практическое занятие' in ss):
                            stype = 'семинар'
                        else:
                            stype = ''


        else:
            ss, ls = pars_func(subjectstring, locationstring, provided_date)
            for s in subjects:
                if s in str(ss):
                    subject = s
                    location = str(ls).replace('\n', '')
                    if 'лекция' in ss:
                        stype = 'лекция'
                    elif ('семинар' in ss) or ('практическое занятие' in ss):
                        stype = 'семинар'
                    else:
                        stype = ''
        l_num += 1
        schedule_list.append(
            {'l_num': l_num, 'time': lessontime, 'subject': subject, 'stype': stype, 'location': location})
    workbook.close()
    return schedule_list